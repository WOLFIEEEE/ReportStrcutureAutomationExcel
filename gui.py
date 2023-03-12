import tkinter as tk
from tkinter import filedialog
import pandas as pd
import openpyxl
from new_try import main as new_try_main
from tkinter import messagebox
import re
import openpyxl
from openpyxl.styles import NamedStyle, Alignment, Font, PatternFill, Border, Side
import openpyxl.utils.cell as column_utils

workbook2 = openpyxl.load_workbook('Boilerplate.xlsx')
Imace_types = ["Critical" , "High" , "Medium" , "Low"]

def addstyles(workbook2):

    cell_style_name = "cell_style"
    if cell_style_name not in workbook2.named_styles:
        cell_style = NamedStyle(name=cell_style_name)
        cell_style.alignment = Alignment(horizontal="left", vertical="center")
        cell_style.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        cell_style.border = Border(left=Side(border_style='thin', color='000000'),
                                   right=Side(border_style='thin', color='000000'),
                                   top=Side(border_style='thin', color='000000'),
                                   bottom=Side(border_style='thin', color='000000'))
        workbook2.add_named_style(cell_style)

    header_cell_style_name = "header_cell_style"
    if header_cell_style_name not in workbook2.named_styles:
        header_cell_style = NamedStyle(name=header_cell_style_name)
        header_cell_style.alignment = Alignment(horizontal="left", vertical="center")
        header_cell_style.fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
        header_cell_style.font = Font(bold=True)
        header_cell_style.border = Border(left=Side(border_style='thin', color='000000'),
                                            right=Side(border_style='thin', color='000000'),
                                            top=Side(border_style='thin', color='000000'),
                                            bottom=Side(border_style='thin', color='000000'))
        workbook2.add_named_style(header_cell_style)

# select the source sheet and the destination sheet
def add_styles_to_pending_cells(sheet , startting_column , ending_column , startting_row , ending_row):
    for row in sheet.iter_rows(min_row=startting_row, max_row=ending_row, min_col=column_utils.column_index_from_string(startting_column), max_col=column_utils.column_index_from_string(ending_column)):
        for cell in row:
            cell.style = "cell_style"

def copy_sheet(source_sheet , destination_sheet):
    for row in source_sheet.iter_rows():
        for cell in row:
            destination_sheet[cell.coordinate].value = cell.value
            destination_sheet.column_dimensions[cell.column_letter].width = source_sheet.column_dimensions[cell.column_letter].width
            adddetialstosheet(destination_sheet)

def adddetialstosheet(destination_sheet):
    for cell in destination_sheet[1]:
        cell.style = "header_cell_style"

def know_the_Columns(sheet, column_name):
    for column in sheet.iter_cols():
        for cell in column:
            # print(cell.value)
            if cell.value == column_name:
                print(cell.column_letter)
                column_index = cell.column_letter

    return column_index


def getnextcolumn(column_index):
    column_index = column_utils.column_index_from_string(column_index)
    column_index = column_index + 1
    column_index = column_utils.get_column_letter(column_index)
    return column_index

def whatcolumntouse(column_name):
    if column_name == 'Level A': return 'H'
    if column_name == 'Level AA': return 'H'
    if column_name == 'Keyboard Navigation': return 'K'
    if column_name == 'Color Contrast': return 'K'
    if column_name == 'Color': return 'K'
    if column_name == 'Zoom': return 'K'
    if column_name == 'HTML Validator': return 'K'
    if column_name == 'Screen Reader': return 'K'
    if column_name == 'Others': return 'K'
    if column_name == 'Critical': return 'I'
    if column_name == 'High': return 'I'
    if column_name == 'Medium': return 'I'
    if column_name == 'Low': return 'I'

    return 'A'

def formattedcellheader(cell_header):
    if cell_header == 'Level A': return 'A'
    if cell_header == 'Level AA': return 'AA'

    return cell_header

def addformulasonrow(sheet , row_number , formula_name):
    heading_row_number = 3;
    starting_column = know_the_Columns(sheet , 'Level A')
    counter = 0
    while counter < 13:
        counter += 1
        cell_header = sheet['{}{}'.format(starting_column, heading_row_number)]
        cell = sheet['{}{}'.format(starting_column, row_number)]
        cell.value = f'=COUNTIFS(\'{formula_name}\'!{whatcolumntouse(cell_header.value)}:{whatcolumntouse(cell_header.value)},"{formattedcellheader(cell_header.value)}")'
        cell.style = "cell_style"
        starting_column = getnextcolumn(starting_column);


def addformulaanddata(execution_sheet , row_name , formula_name , row_number):
    print("inside addformulaanddata")
    cell = None
    column_index = 'A' 
    print("row_numer" , row_number)
    print("column_index" , column_index)
    cell = execution_sheet['{}{}'.format(column_index, row_number)]
    # change the cell value
    cell.value = row_name
    cell.style = "cell_style"
    cell.hyperlink = "#" + formula_name + "!A1"
    addformulasonrow(execution_sheet , row_number , formula_name)

def is_valid_sheet_name(sheet_name):
    """
    Checks if a given text can be used as a valid Excel sheet name according to
    the constraints specified by Microsoft Excel.
    """
    if len(sheet_name) > 31:
        # Sheet name must be no longer than 31 characters
        return False
    elif re.search(r'[\/\*\[\]\:\?\"]', sheet_name):
        # Sheet name cannot contain any of the following characters: / \ * [ ] : ? "
        return False
    elif sheet_name.startswith("'") or sheet_name.endswith("'"):
        # Sheet name cannot begin or end with an apostrophe (')
        return False
    elif sheet_name == ' ' or sheet_name == '':
        # Sheet name cannot be blank or consist solely of spaces
        return False
    else:
        return True


def change_sheet_name(workbook, sheet_name, new_sheet_name):

    return 1
def main_main(Filename , workbook2):
    # Load the workbook
    addstyles(workbook2)
    workbook = openpyxl.load_workbook(filename=Filename)
    execution_sheet = workbook2['Execution Summary']
    source_sheet = workbook2['Sheet_Temp']
    # Get the first sheet
    sheet = workbook.active
    cell = None
    for row in sheet.iter_cols():
        for c in row:
            if c.value == "Sheet Name":
                cell = c
                break
        if cell is not None:
            break

    if cell is None:
        print("Cell with value 'Sheet Name' not found")
    else:
        column_letter = cell.column_letter
        print(column_letter)
        col_index = column_utils.column_index_from_string(column_letter)
        col_index -= 1
        new_column_letter = column_utils.get_column_letter(col_index)
        # Traverse the column and create sheets
        row_number_es = 4
        skip_first_row = True
        for cell in sheet[column_letter]:
            if skip_first_row:
                skip_first_row = False
                continue
            row_number = cell.row
            cell2 = sheet['{}{}'.format(new_column_letter, row_number)]
            
            row_name = cell2.value;
            if(is_valid_sheet_name(cell.value) == False):
                cell.value = show_error_dialog("Sheet Name " + cell.value + " is not valid")
                print("Invalid Sheet Name")
                print(cell.value)
            
            formula_sheet_name = cell.value;
            print("row_name = " + row_name)
            print("formula_sheet_name = " + formula_sheet_name)
            destination_sheet = workbook2.create_sheet(formula_sheet_name)
            copy_sheet(source_sheet, destination_sheet)
            addformulaanddata(execution_sheet, row_name , formula_sheet_name , row_number_es)
            row_number_es = row_number_es + 1

        add_styles_to_pending_cells(execution_sheet , 'B' , know_the_Columns(execution_sheet , 'Level A') , 4 , row_number_es-1)


def getnewcolumnletter(column_letter , deleted_indexes):
    new_column_letter = ord(column_letter)-deleted_indexes
    return new_column_letter

def validate_sheetname(new_value):
    """
    Validates the new sheet name entered by the user and updates the error dialog status label
    """
    if not new_value:
        error_status.set("Sheet name cannot be empty")
        return False
    if len(new_value) > 31:
        error_status.set("Sheet name cannot be longer than 31 characters")
        return False
    if any(c in ['\\', '/', '*', '?', '[', ']', ':'] for c in new_value):
        error_status.set("Sheet name contains invalid characters")
        return False
    if new_value[0] == "'":
        if len(new_value) < 2:
            error_status.set("Sheet name cannot be a single quote")
            return False
        if new_value[-1] != "'":
            error_status.set("Sheet name starting with a quote must end with a quote")
            return False
    error_status.set("") # clear error status if validation passed
    # Check if the new sheet name is valid (you can use your own validation function here)
    return True

def submit_sheetname():
    """
    Gets the value entered by the user and closes the error dialog
    """
    new_sheetname = sheetname_input.get()
    print(new_sheetname)
    
    return new_sheetname

def show_error_dialog(error_text):
    """
    Displays the error dialog with an input field for the new sheet name and a status label
    """
    root_error = tk.Tk()
    root_error.withdraw()

    global error_dialog, error_status
    error_dialog = tk.Toplevel(root_error)
    error_dialog.title("Invalid Sheet Name")
    error_dialog.geometry("300x150")
    
    heading_label = tk.Label(error_dialog, text=" Invalid Sheet Name (" + error_text + ")", fg='red', font=('Arial', 9, 'bold'))
    heading_label.pack(pady=10)


    # heading_label = tk.Label(error_dialog, text=" Invalid Sheet Name (" + error_text + ")", fg='red', font=('Arial', 9, 'bold'))
    # heading_label.pack(pady=10)

    # Create the input field with a predefined constraint
    global sheetname_input
    sheetname_input = tk.Entry(error_dialog, validate='key')
    sheetname_input['validatecommand'] = (sheetname_input.register(validate_sheetname), '%P')
    sheetname_input.pack(pady=10)
    
    # Create the status label
    error_status = tk.StringVar()
    status_label = tk.Label(error_dialog, textvariable=error_status, fg='red')
    status_label.pack(pady=5)
    
    # Create the submit button
    submit_button = tk.Button(error_dialog, text="Submit", command=submit_sheetname)
    submit_button.pack(pady=10)
    new_sheetname = sheetname_input.get()
    
    error_dialog.transient(master=root)
    error_dialog.grab_set()
    error_dialog.wait_visibility()  # Wait until the window is visible
    error_dialog.mainloop()  
    return new_sheetname

def getanarray(selected_platforms_array):
    arr = []
    for i in range(0 , len(selected_platforms_array)):
        if(selected_platforms_array[i] == "Windows"):
            arr.append("Windows")
        if(selected_platforms_array[i] == "macOS"):
            arr.append("macOS")
        if(selected_platforms_array[i] == "Android"):
            arr.append("Android")
        if(selected_platforms_array[i] == "iOS"):
            arr.append("iOS")

    my_strings_only_array = [x for x in arr if isinstance(x, str)]   
    return my_strings_only_array


def open_file():
    # create the BooleanVar objects inside the function
    checkbox_values = {
        "Windows": tk.BooleanVar(),
        "macOS": tk.BooleanVar(),
        "Android": tk.BooleanVar(),
        "iOS": tk.BooleanVar()
    }

    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        # remove the Open File button
        open_button.pack_forget()
        
        # create a label to display the file name
        file_name = file_path.split("/")[-1]
        label = tk.Label(root, text=f"Selected file: {file_name}")
        label.pack(pady=20)

        # create a checkbox for each platform
        for platform, var in checkbox_values.items():
            checkbox = tk.Checkbutton(root, text=platform, variable=var)
            checkbox.pack()

        def process_checkboxes():
            selected_platforms = [platform for platform, var in checkbox_values.items() if var.get()]
            selected_platforms_array = list(selected_platforms)
            print(f"Selected platforms: {selected_platforms}")
            
            # disable the process button and show loading message
            process_button.config(state="disabled", text="Processing...")
            root.update()
            sheet = workbook2["Execution Summary"]
            arr = getanarray(selected_platforms_array)
            new_try_main(arr , sheet)
            
            main_main(file_path , workbook2)
            # enable the process button and show completed message
            process_button.config(state="normal", text="Completed")
            root.update()

            # ask user where to save the new file
            save_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

            # check if user selected a file to save
            if save_file_path:
                # read the processed data
                workbook2.save(save_file_path)
                print(f"File saved at {save_file_path}")

        process_button = tk.Button(root, text="Process Checkboxes", command=process_checkboxes)
        process_button.pack(pady=20, padx=10)

# create the main window
root = tk.Tk()
root.geometry("400x400")

# show_error_dialog("Invalid Sheet Name")

# create a button to open the file dialog
open_button = tk.Button(root, text="Open File", command=open_file)
open_button.pack(pady=20, padx=10)

# run the main loop
root.mainloop()
