import openpyxl
import openpyxl
from openpyxl.styles import NamedStyle, Alignment, Font, PatternFill, Border, Side
import openpyxl.utils.cell as column_utils

# open the workbook



def addstyles(workbook2):

    cell_style_name = "cell_style"
    if cell_style_name not in workbook2.named_styles:
        cell_style = NamedStyle(name=cell_style_name)
        cell_style.alignment = Alignment(horizontal="left", vertical="center")
        cell_style.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        cell_style.border = Border(left=Side(border_style='thin', color='000000'),
                                   right=Side(border_style='thin', color='000000'),
                                   top=Side(border_style='thin', color='000000'),
                                   bottom=Side(border_style='thin', color='000000'))
        workbook2.add_named_style(cell_style)

    header_cell_style_name = "header_cell_style"
    if header_cell_style_name not in workbook2.named_styles:
        header_cell_style = NamedStyle(name=header_cell_style_name)
        header_cell_style.alignment = Alignment(horizontal="left", vertical="center")
        header_cell_style.fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
        header_cell_style.font = Font(bold=True)
        header_cell_style.border = Border(left=Side(border_style='thin', color='000000'),
                                            right=Side(border_style='thin', color='000000'),
                                            top=Side(border_style='thin', color='000000'),
                                            bottom=Side(border_style='thin', color='000000'))
        workbook2.add_named_style(header_cell_style)

# select the source sheet and the destination sheet
def add_styles_to_pending_cells(sheet , startting_column , ending_column , startting_row , ending_row):
    for row in sheet.iter_rows(min_row=startting_row, max_row=ending_row, min_col=column_utils.column_index_from_string(startting_column), max_col=column_utils.column_index_from_string(ending_column)):
        for cell in row:
            cell.style = "cell_style"

def copy_sheet(source_sheet , destination_sheet):
    for row in source_sheet.iter_rows():
        for cell in row:
            destination_sheet[cell.coordinate].value = cell.value
            destination_sheet.column_dimensions[cell.column_letter].width = source_sheet.column_dimensions[cell.column_letter].width
            adddetialstosheet(destination_sheet)

def adddetialstosheet(destination_sheet):
    for cell in destination_sheet[1]:
        cell.style = "header_cell_style"

def know_the_Columns(sheet, column_name):
    for column in sheet.iter_cols():
        for cell in column:
            # print(cell.value)
            if cell.value == column_name:
                print(cell.column_letter)
                column_index = cell.column_letter

    return column_index


def getnextcolumn(column_index):
    column_index = column_utils.column_index_from_string(column_index)
    column_index = column_index + 1
    column_index = column_utils.get_column_letter(column_index)
    return column_index

def whatcolumntouse(column_name):
    if column_name == 'Level A': return 'H'
    if column_name == 'Level AA': return 'H'
    if column_name == 'Keyboard Navigation': return 'K'
    if column_name == 'Color Contrast': return 'K'
    if column_name == 'Color': return 'K'
    if column_name == 'Zoom': return 'K'
    if column_name == 'HTML Validator': return 'K'
    if column_name == 'Screen Reader': return 'K'
    if column_name == 'Others': return 'K'
    if column_name == 'Critical': return 'I'
    if column_name == 'High': return 'I'
    if column_name == 'Medium': return 'I'
    if column_name == 'Low': return 'I'

    return 'A'

def formattedcellheader(cell_header):
    if cell_header == 'Level A': return 'A'
    if cell_header == 'Level AA': return 'AA'

    return cell_header

def addformulasonrow(sheet , row_number , formula_name):
    heading_row_number = 3;
    starting_column = know_the_Columns(sheet , 'Level A')
    counter = 0
    while counter < 13:
        counter += 1
        cell_header = sheet['{}{}'.format(starting_column, heading_row_number)]
        cell = sheet['{}{}'.format(starting_column, row_number)]
        cell.value = f'=COUNTIFS(\'{formula_name}\'!{whatcolumntouse(cell_header.value)}:{whatcolumntouse(cell_header.value)},"{formattedcellheader(cell_header.value)}")'
        cell.style = "cell_style"
        starting_column = getnextcolumn(starting_column);


def addformulaanddata(execution_sheet , row_name , formula_name , row_number):
    print("inside addformulaanddata")
    cell = None
    column_index = 'A' 
    print("row_numer" , row_number)
    print("column_index" , column_index)
    cell = execution_sheet['{}{}'.format(column_index, row_number)]
    # change the cell value
    cell.value = row_name
    cell.style = "cell_style"
    cell.hyperlink = "#" + formula_name + "!A1"
    addformulasonrow(execution_sheet , row_number , formula_name)
    

def main(Filename , workbook2):
    # Load the workbook
    addstyles(workbook2)
    workbook = openpyxl.load_workbook(filename=Filename)
    execution_sheet = workbook2['Execution Summary']
    source_sheet = workbook2['Sheet_Temp']
    # Get the first sheet
    sheet = workbook.active
    cell = None
    for row in sheet.iter_cols():
        for c in row:
            if c.value == "Sheet Name":
                cell = c
                break
        if cell is not None:
            break

    if cell is None:
        print("Cell with value 'Sheet Name' not found")
    else:
        column_letter = cell.column_letter
        print(column_letter)
        col_index = column_utils.column_index_from_string(column_letter)
        col_index -= 1
        new_column_letter = column_utils.get_column_letter(col_index)
        # Traverse the column and create sheets
        row_number_es = 4
        skip_first_row = True
        for cell in sheet[column_letter]:
            if skip_first_row:
                skip_first_row = False
                continue
            row_number = cell.row
            cell2 = sheet['{}{}'.format(new_column_letter, row_number)]
            
            row_name = cell2.value;
            formula_sheet_name = cell.value;
            print("row_name = " + row_name)
            print("formula_sheet_name = " + formula_sheet_name)
            destination_sheet = workbook2.create_sheet(formula_sheet_name)
            copy_sheet(source_sheet, destination_sheet)
            addformulaanddata(execution_sheet, row_name , formula_sheet_name , row_number_es)
            row_number_es = row_number_es + 1

        add_styles_to_pending_cells(execution_sheet , 'B' , know_the_Columns(execution_sheet , 'Level A') , 4 , row_number_es-1)


from openpyxl.utils import column_index_from_string, get_column_letter

def delete_whole_column_letter(columns):
    workbook2 = openpyxl.load_workbook('Boilerplate.xlsx')
    sheet = workbook2['Execution Summary']
    deleted_indexes = []
    for col in columns:
        col_index = column_index_from_string(col)
        sheet.delete_cols(col_index, 1)
        deleted_indexes.append(col_index)
    # Update remaining column letters
    for cell in sheet.iter_cols(min_row=1, min_col=min(deleted_indexes), max_col=max(deleted_indexes)):
        new_col_index = column_index_from_string(cell[0].column_letter) - len(deleted_indexes)
        print(f"new_col_index: {new_col_index}")  # add this line
        new_col_letter = get_column_letter(new_col_index)
        for c in cell:
            c.column_letter = new_col_letter

# def delete_whole_column_letter(arr):
#     workbook2 = openpyxl.load_workbook('Boilerplate.xlsx')
#     execution_sheet = workbook2['Execution Summary']
#     for column in arr:
#         if(column == 'Windows'):
#             execution_sheet.delete_cols(column_utils.column_index_from_string('B'))
#         if(column == 'Android'):
#             execution_sheet.delete_cols(column_utils.column_index_from_string('D'))
#         if(column == 'iOS'):
#             execution_sheet.delete_cols(column_utils.column_index_from_string())
#         if(column == 'macOS'):
#             column = 'C'
        
#     workbook2.save('Boilerplate.xlsx')
    
# Filename = 'try123456.xlsx'
# main(Filename)
# Save the workbook

