import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


Conf_levels = ['Level A' , 'Level AA']
Conf_Title = "Conformance Level Wise Distribution"
Issue_types = ["Keyboard Navigation" , "Color Contrast" , "Color" , "Zoom", "HTML validator", "Screen Reader" , "Other A11y"]
Issue_Title = "Accessibility Issue Type Distribution"
Imace_types = ["Critical" , "High" , "Medium" , "Low"]
Imace_Title = "Imace Wise Distribution"
Execution_Title = "Execution Status"

fill = PatternFill(start_color='FFB4C6E7', end_color='FFB4C6E7', fill_type='solid')
border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))

def add_first_column(column_name , worksheet):
    # Merge cells and set value
    worksheet.merge_cells('A2:A3')
    worksheet['A2'] = column_name
    # Apply font and alignment
    font = Font(bold=True)
    alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    worksheet['A2'].font = font
    worksheet['A2'].alignment = alignment
    worksheet['A2'].fill = fill
    worksheet['A2'].border = border
    worksheet.column_dimensions['A'].width = 49


def get_column_width(column_name):
    if column_name == "Level A" or column_name == "Level AA":
        return 20
    if column_name == "Keyboard Navigation" or column_name == "Color Contrast" or column_name == "HTML validator" or column_name == "Screen Reader" :
        return 22
    if column_name == "Critical" or column_name == "High" or column_name == "Medium" or column_name == "Low":
        return 12
    if column_name == "Color" or column_name == "Zoom" or column_name == "Other A11y":
        return 11
    else:
        return 15
    

def merge_cells_and_set_values(column_letter, worksheet, text_arr , heading_text):

    print("Array Revecied")
    #print all the values of text arr 
    for i in range(len(text_arr)):
        print(text_arr[i])
    # get the size of the text array
    size = len(text_arr)
    new_col_index = openpyxl.utils.column_index_from_string(column_letter) + size - 1
    new_col = openpyxl.utils.get_column_letter(new_col_index)
    merge_range = f"{column_letter}2:{new_col}2"
    worksheet.merge_cells(merge_range)

    # set the text and bold font for the merged cell
    worksheet[f"{column_letter}2"] = heading_text
    worksheet[f"{column_letter}2"].font = Font(bold=True)
    worksheet[f"{column_letter}2"].alignment = Alignment(horizontal='center')
    for row in worksheet[merge_range]:
        for cell in row:
            cell.fill = fill
            cell.border = border

    # set the text for the 6th row based on the array
    for i in range(size):
        if text_arr[i] is None:
            text_arr[i] = ""
        worksheet[f"{column_letter}3"] = text_arr[i]
        worksheet[f"{column_letter}3"].font = Font(bold=True)
        worksheet[f"{column_letter}3"].alignment = Alignment(horizontal='center')
        worksheet[f"{column_letter}3"].fill = fill
        worksheet[f"{column_letter}3"].border = border
        worksheet.column_dimensions[column_letter].width = get_column_width(text_arr[i])

        new_col_index = openpyxl.utils.column_index_from_string(column_letter) + 1;
        column_letter = openpyxl.utils.get_column_letter(new_col_index)

    return column_letter

# create a new workbook
def main(arr , worksheet):
    add_first_column("Pages / Flows" , worksheet)
    column_letter = 'B'
    column_letter =  merge_cells_and_set_values(column_letter, worksheet, arr, Execution_Title)
    column_letter =  merge_cells_and_set_values(column_letter, worksheet, Conf_levels, Conf_Title)
    column_letter =  merge_cells_and_set_values(column_letter, worksheet, Issue_types, Issue_Title)
    column_letter =  merge_cells_and_set_values(column_letter, worksheet, Imace_types, Imace_Title)
